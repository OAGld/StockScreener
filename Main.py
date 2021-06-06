from Scraper import Scrape_Data, Scrape_Yahoo_Finance
import openpyxl
from openpyxl import load_workbook
import configparser

#Open the TickerList excel document
TickerList = load_workbook(filename="TickerList.xlsx")
TickerSheet = TickerList.active

#Create a list for the Scrape_Data instances
Comp_Data = []

#Create a new excel document to write new data to
CompaniesForAnalysis = openpyxl.Workbook()


def main():
    #CompaniesForAnalysis = openpyxl.Workbook()
    Counter = 1
    Index = 0
    parser = configparser.ConfigParser()
    parser.read("StockCriteria.txt")

    while TickerSheet.cell(row=Counter, column=1).value != None:
        Comp_Data.append(Scrape_Data(TickerSheet.cell(row=Counter, column=1).value))
        Counter += 1

    for a in range(len(Comp_Data)):
        Dividend = Calc_Dividend(a)

        if Comp_Data[a].Current_Ratio != None and Comp_Data[a].TrailingYield != None and Comp_Data[a].CurrentYield != None and Comp_Data[a].Profit_Margin != None and Comp_Data[a].RoA != None and Comp_Data[a].RoE != None:
            if (Comp_Data[a].Current_Ratio >= float(parser.get("criteria", "Current_Ratio"))) and (Comp_Data[a].TrailingYield >= float(parser.get("criteria", "TrailingYield")) or Comp_Data[a].CurrentYield >= float(parser.get("criteria", "CurrentYield"))) and Comp_Data[a].Profit_Margin >= float(parser.get("criteria", "Profit_Margin")) and Comp_Data[a].RoA >= float(parser.get("criteria", "RoA")) and Comp_Data[a].RoE >= float(parser.get("criteria", "RoE")):
                try:
                    CompaniesForAnalysis.create_sheet(index=Index, title=Comp_Data[a].Comp_Name)
                    sheet = CompaniesForAnalysis.get_sheet_by_name(Comp_Data[a].Comp_Name)
                    sheet.column_dimensions['A'].width = 30
                    sheet.title = Comp_Data[a].Comp_Name
                except Exception as e:
                    print("Problems with excel worksheet")
                    print(e)

                try:
                    #Insert all values
                    sheet['A1'] = 'Ticker: '
                    sheet['B1'] = Comp_Data[a].Comp_Name
                    sheet['A2'] = 'Profit Margin: '
                    sheet['B2'] = str(Comp_Data[a].Profit_Margin) + '%'
                    sheet['A3'] = 'Return on Assets: '
                    sheet['B3'] = str(Comp_Data[a].RoA) + '%'
                    sheet['A4'] = 'Return on Equity: '
                    sheet['B4'] = str(Comp_Data[a].RoE) + '%'
                    sheet['A5'] = 'Current Ratio: '
                    sheet['B5'] = str(Comp_Data[a].Current_Ratio)
                    sheet['A6'] = 'Debt-to-Equity Ratio: '
                    sheet['B6'] = str(round(float(Comp_Data[a].D_to_E_Ratio)/100, 2))
                    sheet['A7'] = 'Shares Outstanding: '
                    sheet['B7'] = Comp_Data[a].Shares_Outstanding
                    sheet['A8'] = 'Payout Ratio: '
                    sheet['B8'] = Comp_Data[a].Payout_Ratio
                    sheet['A9'] = 'Forward Annual Dividend Yield: '
                    sheet['B9'] = str(Comp_Data[a].CurrentYield) + '%'
                    sheet['A10'] = 'Trailing Annual Dividend Yield: '
                    sheet['B10'] = str(Comp_Data[a].TrailingYield) + '%'
                    sheet['A11'] = 'Average dividend growth: '
                    sheet['B11'] = Dividend.DividendGrowth
                    sheet['D1'] = 'Year'
                    sheet['E1'] = 'Dividend'

                    for i in range(len(Comp_Data[a].List)):
                        sheet['D' + str(i + 2)] = Comp_Data[a].List[i][0]
                        sheet['E' + str(i + 2)] = Comp_Data[a].List[i][1]
                except Exception as e:
                    print("Problems inserting values in excel sheet")
                    print(e)

                Index += 1

    #CompaniesForAnalysis.remove_sheet(CompaniesForAnalysis.get_sheet_by_name('Sheet'))
    try:
        #del CompaniesForAnalysis['Sheet']
        CompaniesForAnalysis.save('CompaniesForAnalysis.xlsx')
    except Exception as e:
        print("Problems deleting sheet or saving excel sheet")
        print(e)

class Calc_Dividend:

    def __init__(self, Instance):
        self.DividendGrowth = None
        self.Instance = Instance
        Calc_Dividend.Dividend_Growth(self)

    def Dividend_Growth(self):
        try:
            if len(Comp_Data[self.Instance].List) >= 5:
                Sum = 0
                Annual_Growth_Rate = []
                for a in range(len(Comp_Data[self.Instance].List)):
                    if a > 0:
                        Annual_Growth_Rate.append((float(Comp_Data[self.Instance].List[a-1][1])/float(Comp_Data[self.Instance].List[a][1]))-1)

                for a in range(len(Annual_Growth_Rate)):
                    Sum += Annual_Growth_Rate[a]

                self.DividendGrowth = str(round((Sum/float(len(Annual_Growth_Rate))*100), 2)) + '%'
            else:
                self.DividendGrowth = "Too few"
        except Exception as e:
            print("Problems calculating dividend growth.")
            print(e)

main()
